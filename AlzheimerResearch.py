import pyrealsense2 as rs
import numpy as np
import cv2
import os
import tkinter as tk
from tkinter import messagebox
from datetime import datetime
import sys
from openpyxl import Workbook
import serial
import serial.tools.list_ports
import time
import threading
import queue

global_save_path = None  # 全局變量來存儲 save_path

# ====輔助函數 - 發送指令給Arduino====#

def find_arduino_ports():
    ports = serial.tools.list_ports.comports() # 列出所有串接埠
    arduino_ports = [port.device for port in ports if 'Arduino' in port.description or 'CH340' in port.description]
    return arduino_ports

def flash_white_light_3_times(port):
    try:
        with serial.Serial(port, 9600, timeout=2) as ser: # 開啟 serial port
            time.sleep(2)  # 給 Arduino 一點時間來重置和準備接收數據，不能刪，不能小於2！！！
            ser.write(b'flash_white_light_3_times\n')
    except serial.SerialException as e:
        print(f"無法開啟串列埠 {port}: {e}")

def LED_cycle_3_times(port):
    try:
        with serial.Serial(port, 9600, timeout=2) as ser:
            time.sleep(2)  # 給 Arduino 一點時間來重置和準備接收數據，不能刪，不能小於2！！！
            ser.write(b'LED_cycle_3_times\n')
    except serial.SerialException as e:
        print(f"無法開啟串列埠 {port}: {e}")

# ====輔助函數 - UI界面====#

def show_error(message):
    messagebox.showerror("Error", message)

def create_label_entry(window, text, row, font=('microsoft yahei', 14, 'bold'), pady=(10, 10), padx=(10, 10), ipady=10):
    label = tk.Label(window, text=text, font=font) # 建立文字label
    label.grid(row=row, column=0) # 指定放置位置
    entry = tk.Entry(window) # 建立單行輸入框
    entry.grid(row=row, column=1, padx=padx, pady=pady, ipady=ipady) # 指定放置位置
    return label, entry

def create_button(window, text, command, row, column):
    """創建按鈕的統一函數。"""
    button = tk.Button(window, text=text, command=command, font=('microsoft yahei', 14, 'bold'))
    button.grid(row=row, column=column)

def check_entry_value(input_name_entry, input_date_entry, input_sex_entry, input_birth_entry):
    input_name_value = input_name_entry.get()
    input_date_value = input_date_entry.get()
    input_sex_value = input_sex_entry.get().upper()  # 將輸入轉為大寫統一處理
    input_birth_value = input_birth_entry.get()

    cwd = os.getcwd()
    save_folder = os.path.join(cwd, 'saved_file/')
    new_folder_path = os.path.join(save_folder, input_name_value)

    error_flag = 0

    # 檢查輸入是否正確
    if os.path.exists(new_folder_path):
        show_error("檔名輸入錯誤，輸入相同檔名")
        error_flag = 1

    try:
        datetime.strptime(input_date_value, '%Y%m%d')
    except ValueError:
        input_date_entry.delete(0, tk.END)
        show_error("拍攝日期格式不正確，請重新輸入")
        error_flag = 1

    try:
        datetime.strptime(input_birth_value, '%Y%m%d')
    except ValueError:
        input_birth_entry.delete(0, tk.END)
        show_error("生日格式不正確，請重新輸入")
        error_flag = 1

    if input_sex_value not in ["F", "M"]:
        input_sex_entry.delete(0, tk.END)
        show_error("您输入的性别不正確，請重新输入")
        error_flag = 1


    return error_flag, new_folder_path

def save_excel(input_name_entry,input_date_entry,input_sex_entry,input_birth_entry):
    input_name_value = input_name_entry.get()
    input_date_value = input_date_entry.get()
    input_sex_value = input_sex_entry.get().upper()  # 將輸入轉為大寫統一處理
    input_birth_value = input_birth_entry.get()
    
    wb = Workbook()  # creates a workbook object.
    ws_x = wb.create_sheet("病人資料", 0)
    ws_x.cell(row=2, column=1).value = input_name_value
    ws_x.cell(row=2, column=2).value = input_date_value
    ws_x.cell(row=2, column=3).value = input_sex_value
    ws_x.cell(row=2, column=4).value = input_birth_value

    ws_x.cell(row=1, column=1).value = '編號'
    ws_x.cell(row=1, column=2).value = '拍攝日期'
    ws_x.cell(row=1, column=3).value = '性別'
    ws_x.cell(row=1, column=4).value = '生日'

    wb.save(global_save_path+'/' + '病人資料.xlsx')  # save to excel file.

def send_input(window, input_name_entry, input_date_entry, input_sex_entry, input_birth_entry, is_debug=False):
    print('Button clicked!')
    global global_save_path

    if is_debug:
        input_name_entry.insert(0, f"test_{datetime.now().strftime('%Y%m%d%H%M%S')}") 
        input_date_entry.insert(0, '20210901')
        input_sex_entry.insert(0, 'M')
        input_birth_entry.insert(0, '19900101')

    error_flag, new_folder_path = check_entry_value(input_name_entry, input_date_entry, input_sex_entry, input_birth_entry)
    if error_flag==0:
        # 關閉UI界面
        os.makedirs(new_folder_path, exist_ok=True)
        global_save_path = new_folder_path  # 存儲 save_path 到全局變量
        save_excel(input_name_entry, input_date_entry, input_sex_entry, input_birth_entry)
        window.destroy()
        
def clear_input(input_name_entry, input_date_entry, input_sex_entry, input_birth_entry):
    input_name_entry.delete(0, tk.END)
    input_date_entry.delete(0, tk.END)
    input_sex_entry.delete(0, tk.END)
    input_birth_entry.delete(0, tk.END)

# ====主函數==== #

def UI_window(is_debug=True):
    window = tk.Tk()
    window.title('請輸入病人資訊')
    window_width = window.winfo_screenwidth() # 1440
    window_height = window.winfo_screenheight() # 900
    width = 600
    height = 400
    left = int((window_width - width)/2)       # 計算左上 x 座標
    top = int((window_height - height)/2)      # 計算左上 y 座標
    window.geometry(f'{width}x{height}+{left}+{top}') # 寬、高、左上角座標
    window.resizable(False, False)

    # 創建標籤和輸入框
    _, input_name_entry = create_label_entry(window, "編號", 0)
    _, input_date_entry = create_label_entry(window, "拍攝日期(YYYYMMDD)", 1)
    _, input_sex_entry = create_label_entry(window, "性別:(M/F)", 2)
    _, input_birth_entry = create_label_entry(window, "生日(YYYYMMDD)", 3)

     # 使用函數創建按鈕
    create_button(window, '送出資料', lambda: send_input(window, input_name_entry, input_date_entry, input_sex_entry, input_birth_entry, is_debug=is_debug), 5, 1)
    create_button(window, '關閉', lambda: sys.exit(), 5, 2)
    create_button(window, '重新輸入資料', lambda: clear_input(input_name_entry, input_date_entry, input_sex_entry, input_birth_entry), 5, 3)

    # 開始事件循環
    window.mainloop()

def trigger_whight_light():
    print('Trigger light, pass now.')
    pass




def show_frame(pipeline):
    # 讀取一幀影像
    frames = pipeline.wait_for_frames()
    color_frame = frames.get_color_frame()
    depth_frame = frames.get_depth_frame()

    if not color_frame or not depth_frame:
        return None
    color_image = np.asanyarray(color_frame.get_data())
    color_image = cv2.transpose(color_image)
    original_color_image = color_image.copy()
    depth_image = np.asanyarray(depth_frame.get_data())
    depth_image = cv2.transpose(depth_image)

    # 繪製需對齊的正方型
    x_center = int(color_image.shape[1] / 2)
    y_center = int(color_image.shape[0] / 2)
    square_size = 80
    cv2.rectangle(color_image, (x_center-square_size, y_center-square_size), (x_center+square_size, y_center+square_size), (0, 255, 0), 2)
    
    have_face = detect_face_distance(face_cascade, color_image, depth_image, x_center, y_center, square_size)

    # 顯示彩色影像
    color_image = cv2.resize(color_image, (0, 0), fx=0.75, fy=0.75)
    # cv2.imshow('Color Image', color_image)

    return original_color_image, color_image, depth_image, have_face



def detect_face_distance(face_cascade, color_image, depth_image, x_center, y_center, square_size):

    have_face = False

    # 偵測人臉
    gray_image = cv2.cvtColor(color_image, cv2.COLOR_BGR2GRAY)
    faces = face_cascade.detectMultiScale(gray_image, scaleFactor=1.03, minNeighbors=7, minSize=(100, 100))

    # 若有大於一個人臉，則選擇最大的人臉
    if len(faces) > 1:
        max_area = 0
        for (x, y, w, h) in faces:
            if w * h > max_area:
                max_area = w * h
                max_face = (x, y, w, h)
        faces = [max_face]

    # 針對人臉繪製方框
    for (x, y, w, h) in faces:
        x_center_face, y_center_face = int(x + w / 2), int(y + h / 2)
        cv2.circle(color_image, (x_center_face, y_center_face), 5, (0, 0, 255), 2)
        
        depth_roi = depth_image[y:y + h, x:x + w] / 1000
        valid_mask = (depth_roi > 0.1) & (depth_roi < 1)
        valid_distances = depth_roi[valid_mask]
        ave_distance = np.mean(valid_distances) if valid_distances.size > 0 else 0
        
        # 設定方框顏色，綠色代表距離不合適，紅色代表距離合適
        color = (0, 255, 0)
        if 0.35 <= ave_distance <= 0.5:
            if x_center - square_size < x_center_face < x_center + square_size and y_center - square_size < y_center_face < y_center + square_size:
                color = (0, 0, 255)
                have_face = True

        # 繪製方框及文字
        text = f"Distance: {ave_distance:.2f} m"
        cv2.rectangle(color_image, (x, y), (x + w, y + h), color, 2)
        cv2.putText(color_image, text, (x, y - 10), cv2.FONT_HERSHEY_SIMPLEX, 0.5, color, 2)

    return have_face

# def save_frame(color_image, save_path, index):
#     # 儲存影像
#     cv2.imwrite(f'{save_path}/{index}.png', color_image)


class ImageSaver:
    def __init__(self, num_worker_threads, global_save_path):
        self.image_queue = queue.Queue()
        self.global_save_path = global_save_path
        self.threads = []
        for _ in range(num_worker_threads):
            t = threading.Thread(target=self.save_worker)
            t.start()
            self.threads.append(t)

    def save_worker(self):
        while True:
            item = self.image_queue.get()
            if item is None:
                break
            original_color_image, saved_images_count = item
            self.save_frame(original_color_image, saved_images_count)
            self.image_queue.task_done()

    def save_frame(self, original_color_image, saved_images_count):
        cv2.imwrite(f'{self.global_save_path}/{saved_images_count}.png', original_color_image)
        pass

    def add_image(self, original_color_image, saved_images_count):
        self.image_queue.put((original_color_image, saved_images_count))

    def stop(self):
        for _ in range(len(self.threads)):
            self.image_queue.put(None)
        for t in self.threads:
            t.join()


class RealsenseCamera:
    def __init__(self):
        self.pipeline = rs.pipeline()
        self.config = rs.config()
        self.height = 720
        self.width = 1280
        self.config.enable_stream(rs.stream.depth, self.width, self.height, rs.format.z16, 30)
        self.config.enable_stream(rs.stream.color, self.width, self.height, rs.format.bgr8, 30)
        self.pipeline.start(self.config)


    def get_frame(self):
        # 讀取一幀影像
        self.frames = self.pipeline.wait_for_frames()
        color_frame = self.frames.get_color_frame()
        depth_frame = self.frames.get_depth_frame()

        if not color_frame or not depth_frame:
            return None
        color_image = np.asanyarray(color_frame.get_data())
        color_image = cv2.transpose(color_image)
        original_color_image = color_image.copy()
        depth_image = np.asanyarray(depth_frame.get_data())
        depth_image = cv2.transpose(depth_image)

        # 繪製需對齊的正方型
        x_center = int(color_image.shape[1] / 2)
        y_center = int(color_image.shape[0] / 2)
        square_size = 80
        cv2.rectangle(color_image, (x_center-square_size, y_center-square_size), (x_center+square_size, y_center+square_size), (0, 255, 0), 2)
        
        have_face = detect_face_distance(face_cascade, color_image, depth_image, x_center, y_center, square_size)

        # 顯示彩色影像
        color_image = cv2.resize(color_image, (0, 0), fx=0.75, fy=0.75)

        return original_color_image, color_image, depth_image, have_face


if __name__ == '__main__':

    # 輸入個人資料
    UI_window()

    # 模型初始化
    face_cascade = cv2.CascadeClassifier(os.path.join(os.path.dirname(os.path.abspath(__file__)), 'haarcascade_frontalface_default.xml'))

    # # 相機初始化
    # pipeline = rs.pipeline()
    # config = rs.config()

    # height = 720
    # width = 1280

    # config.enable_stream(rs.stream.depth, width, height, rs.format.z16, 30)
    # config.enable_stream(rs.stream.color, width, height, rs.format.bgr8, 30)
    # pipeline.start(config)

    rs_camera = RealsenseCamera()

    # Arduino初始化
    ports = find_arduino_ports()

    print("找到以下可能是 Arduino 的串列埠：")
    for port in ports:
        print(port)

    if ports is None:
        raise ValueError("找不到 'Arduino' 的串列埠。")
    

    try:
        taking_images = False
        have_face = False
        saved_images_count = 0
        image_saver = ImageSaver(8, global_save_path)


        while True:
    
            key = cv2.waitKey(1) & 0xFF
            if key == ord('q') or key == ord('Q'):
                print('ESC key pressed.')
                break
            if key == ord('s') or key == ord('S'):
                if have_face:                    
                    print('Face detected, begin to save images.')
                    taking_images = True
                    flash_white_light_3_times(ports[0])
                    time.sleep(5)
                    LED_cycle_3_times(ports[0])
                    start_time = time.time()
                else:
                    print('No face detected or position not good.')
                    continue
            

            original_color_image, color_image, depth_image, have_face = rs_camera.get_frame()
            cv2.imshow('Color Image', color_image)

            if taking_images:
                # save_frame(original_color_image, global_save_path, saved_images_count)
                image_saver.add_image(original_color_image, saved_images_count)
                print('Save image:', f'{global_save_path}/{saved_images_count}.png')
                saved_images_count += 1

            if saved_images_count == 1200:
                print('All images saved.')
                print('Time used:', time.time() - start_time)
                taking_images = False
                saved_images_count = 0
                image_saver.stop()
                continue
                

    finally:
        cv2.destroyAllWindows()
        print('Close the camera and exit.')
